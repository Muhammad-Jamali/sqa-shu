import openpyxl

workbook = openpyxl.load_workbook("..//excel//data.xlsx")
sheet = workbook["Sheet1"]

total_rows = sheet.max_row
total_cols = sheet.max_column

#print(total_rows,total_cols)

#print(sheet.cell(row=1,column=1).value)
#print(sheet.cell(row=sheet.max_row,column=sheet.max_column).value)

for rows in range(1,total_rows+1):
    for cols in range(1,total_cols+1):
        print(sheet.cell(row= rows, column=cols).value, end=" ")
    print()