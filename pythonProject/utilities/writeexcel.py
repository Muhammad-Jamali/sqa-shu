import openpyxl

workbook = openpyxl.load_workbook("..//excel//data.xlsx")
sheet = workbook["Sheet1"]

sheet.cell(row=5,column=1).value = "sample4@gmail.com"
sheet.cell(row=5,column=2).value = "44444"

workbook.save("..//excel//data.xlsx")


total_rows = sheet.max_row
total_cols = sheet.max_column

for rows in range(1,total_rows+1):
    for cols in range(1,total_cols+1):
        print(sheet.cell(row= rows, column=cols).value, end=" ")
    print()