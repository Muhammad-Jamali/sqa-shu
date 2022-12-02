from configparser import ConfigParser
from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
import pytest
import openpyxl

def setup_function():
    global driver
    driver = webdriver.Chrome(executable_path=ChromeDriverManager().install())
    driver.get(readConfig("basic","url"))
    driver.maximize_window()

def teardown_functin():
    driver.quit()

def get_data():
    workbook = openpyxl.load_workbook("..//excel//data.xlsx")
    sheet = workbook["Sheet1"]
    totalrows = sheet.max_row
    totalcols = sheet.max_column
    mainList = []

    for i in range(2,totalrows+1):
        dataList = []
        for j in range(1, totalcols+1):
           data = sheet.cell(row=i,column=j).value
           dataList.insert(j,data)
        mainList.insert(i,dataList)
    return mainList

def readConfig(section,key):
    config = ConfigParser()
    config.read("config.ini")
    return config.get(section,key)

@pytest.mark.parametrize("username,password", get_data())
def test_dologin(username, password):
    driver.find_element(By.XPATH,readConfig("locator","username")).send_keys(username)
    driver.find_element(By.XPATH,readConfig("locator","password")).send_keys(username)
